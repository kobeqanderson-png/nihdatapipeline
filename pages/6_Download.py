"""
Page 6: Download Processed Data
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from pathlib import Path
import sys
from scipy import stats as scipy_stats
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

try:
    from src.branding import brand_label
except ModuleNotFoundError:
    def brand_label(default: str = "kobeanderson-png") -> str:
        return default
from src.navigation import apply_global_chrome, render_top_navigation

st.set_page_config(page_title="Download", page_icon=None, layout="wide")

apply_global_chrome()
render_top_navigation("Download")

st.title("Download Processed Data")

if st.session_state.df_processed is None:
    st.warning("No processed data available. Please run the **Process Data** pipeline first.")
    st.stop()

df_processed = st.session_state.df_processed

st.markdown("""
Export your processed data in multiple formats with detailed summary statistics.
""")

st.divider()

# Helper functions
def parse_animal_number(value) -> float:
    """Extract animal ID numbers from numeric values or labels."""
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip()
    if not text:
        return np.nan
    lower_text = text.lower()
    import re
    prefixed_match = re.search(r"(?:rat|subject|animal)\s*[-_#:]?\s*(\d+(?:\.\d+)?)", lower_text)
    if prefixed_match:
        return float(prefixed_match.group(1))
    numeric_only_match = re.fullmatch(r"\d+(?:\.\d+)?", lower_text)
    if numeric_only_match:
        return float(numeric_only_match.group(0))
    match = re.search(r"\d+(?:\.\d+)?", text)
    if match:
        return float(match.group(0))
    return np.nan

def parse_animal_number_series(series: pd.Series) -> pd.Series:
    """Vectorized animal-number parsing."""
    return series.apply(parse_animal_number)

def infer_animal_column(df: pd.DataFrame):
    """Infer likely animal identifier column name."""
    normalized = {col.lower().strip(): col for col in df.columns}
    preferred = ['animal #', 'animal', 'animal number', 'animal_number', 'animal id', 'animal_id']
    for key in preferred:
        if key in normalized:
            return normalized[key]
    for col in df.columns:
        if 'animal' in col.lower():
            return col
    return None

def ttest_for_groups(df: pd.DataFrame, value_col: str, group_col: str = "Sex"):
    """Compute Welch's t-test between Male and Female groups."""
    male_data = df[df[group_col] == 'Male'][value_col].dropna()
    female_data = df[df[group_col] == 'Female'][value_col].dropna()
    if len(male_data) > 1 and len(female_data) > 1:
        t_stat, p_value = scipy_stats.ttest_ind(male_data, female_data, equal_var=False)
        return t_stat, p_value, len(male_data), len(female_data)
    return np.nan, np.nan, len(male_data), len(female_data)

def sem(series: pd.Series) -> float:
    """Calculate SEM = std(N) / sqrt(N)."""
    values = series.dropna()
    n = len(values)
    if n < 2:
        return np.nan
    return values.std(ddof=1) / np.sqrt(n)

def autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    """Autosize worksheet columns based on cell text length."""
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_width), max_width)

def build_excel_export(
    df_processed: pd.DataFrame,
    animal_col: str = None,
    threshold: float = 16,
    rebuild_sex_labels: bool = True,
    include_branding: bool = False,
) -> bytes:
    """Build Excel sheet with animal-ordered data and summary stats."""
    excel_buffer = io.BytesIO()

    animal_col = animal_col if animal_col in df_processed.columns else infer_animal_column(df_processed)
    export_df = df_processed.copy()

    if animal_col and rebuild_sex_labels:
        animal_num = parse_animal_number_series(export_df[animal_col])
        export_df['Sex'] = np.where(
            animal_num <= threshold,
            'Male',
            np.where(animal_num > threshold, 'Female', 'Unclassified')
        )
    elif 'Sex' not in export_df.columns:
        export_df['Sex'] = 'Unclassified'

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        sheet_name = 'Summary_By_Animal'
        header_font = Font(bold=True)
        section_font = Font(bold=True, size=12)
        highlight_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
        section_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        table_header_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        ordered_df = export_df.copy()
        low_count = 0
        has_gap = False
        
        if animal_col:
            ordered_df['_animal_num_sort'] = parse_animal_number_series(ordered_df[animal_col])
            low = ordered_df[ordered_df['_animal_num_sort'].notna() & (ordered_df['_animal_num_sort'] <= threshold)]
            high = ordered_df[ordered_df['_animal_num_sort'].notna() & (ordered_df['_animal_num_sort'] > threshold)]
            unknown = ordered_df[ordered_df['_animal_num_sort'].isna()]

            low_count = len(low)
            has_gap = len(low) > 0 and len(high) > 0

            pieces = [low]
            if has_gap:
                spacer = pd.DataFrame([{col: np.nan for col in ordered_df.columns}])
                pieces.append(spacer)
            pieces.extend([high, unknown])
            ordered_df = pd.concat(pieces, ignore_index=True)
            ordered_df = ordered_df.drop(columns=['_animal_num_sort'])

        ordered_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]

        if include_branding:
            try:
                ws.oddFooter.right.text = brand_label()
                ws.oddFooter.right.size = 8
            except Exception:
                pass

        ws.cell(row=1, column=1, value='Summary Ordered by Animal Number').font = Font(bold=True, size=13)
        ws.cell(row=3, column=1, value='Animal-Ordered Processed Data').font = section_font
        ws.cell(row=3, column=1).fill = section_fill

        total_cols = len(ordered_df.columns)
        header_row = 3
        data_start = 4

        for c in range(1, total_cols + 1):
            ws.cell(row=header_row, column=c).font = header_font
            ws.cell(row=header_row, column=c).fill = table_header_fill
            ws.cell(row=header_row, column=c).border = thin_border

        if animal_col and has_gap:
            spacer_row = data_start + low_count
            ws.cell(
                row=spacer_row,
                column=1,
                value=f'--- Spacer: animals > {int(threshold) if float(threshold).is_integer() else threshold} begin below ---'
            ).font = header_font
            ws.cell(row=spacer_row, column=1).fill = section_fill

        for r in range(data_start, data_start + len(ordered_df)):
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).border = thin_border

        # Summary stats table
        summary_title_row = data_start + len(ordered_df) + 2
        summary_header_row = summary_title_row + 1
        summary_data_start = summary_header_row + 1

        summary_headers = [
            'Variable', 'Male N', 'Female N', 'Male Mean', 'Female Mean', 'Male SEM', 'Female SEM',
            't-stat', 'p-value', 'p<0.005'
        ]

        ws.cell(row=summary_title_row, column=1, value='Sex Summary Table (Welch t-test)').font = Font(bold=True, size=12)
        ws.cell(row=summary_title_row + 1, column=1, value='SEM = stdev/sqrt(n)').font = header_font

        for i, col_title in enumerate(summary_headers, start=1):
            hcell = ws.cell(row=summary_header_row, column=i, value=col_title)
            hcell.font = header_font
            hcell.fill = table_header_fill
            hcell.border = thin_border

        numeric_cols = [
            col for col in export_df.select_dtypes(include=['number']).columns
            if col != animal_col
        ]

        for i, col_name in enumerate(numeric_cols):
            row = summary_data_start + i
            male_vals = export_df[export_df['Sex'] == 'Male'][col_name].dropna()
            female_vals = export_df[export_df['Sex'] == 'Female'][col_name].dropna()
            t_stat, p_value, n_male, n_female = ttest_for_groups(export_df, col_name)
            is_highlight = not pd.isna(p_value) and p_value < 0.005

            male_mean = male_vals.mean() if len(male_vals) > 0 else np.nan
            female_mean = female_vals.mean() if len(female_vals) > 0 else np.nan
            male_sem = sem(male_vals)
            female_sem = sem(female_vals)

            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=n_male)
            ws.cell(row=row, column=3, value=n_female)
            ws.cell(row=row, column=4, value=round(float(male_mean), 6) if not pd.isna(male_mean) else np.nan)
            ws.cell(row=row, column=5, value=round(float(female_mean), 6) if not pd.isna(female_mean) else np.nan)
            ws.cell(row=row, column=6, value=round(float(male_sem), 6) if not pd.isna(male_sem) else np.nan)
            ws.cell(row=row, column=7, value=round(float(female_sem), 6) if not pd.isna(female_sem) else np.nan)
            ws.cell(row=row, column=8, value=round(float(t_stat), 6) if not pd.isna(t_stat) else np.nan)
            ws.cell(row=row, column=9, value=round(float(p_value), 6) if not pd.isna(p_value) else np.nan)
            ws.cell(row=row, column=10, value='YES' if is_highlight else 'NO')

            for col_num in range(1, len(summary_headers) + 1):
                ws.cell(row=row, column=col_num).border = thin_border

            if is_highlight:
                for col_num in range(1, len(summary_headers) + 1):
                    ws.cell(row=row, column=col_num).fill = highlight_fill

        ws.freeze_panes = 'A4'
        if len(numeric_cols) > 0:
            ws.auto_filter.ref = (
                f"A{summary_header_row}:{get_column_letter(len(summary_headers))}{summary_data_start + len(numeric_cols) - 1}"
            )
        autosize_columns(ws)

    return excel_buffer.getvalue()

# Download options
st.header("Export Formats")

col1, col2 = st.columns(2)

with col1:
    st.subheader("CSV Format")
    st.caption("Simple comma-separated values file")
    
    csv_buffer = io.StringIO()
    df_processed.to_csv(csv_buffer, index=False)
    csv_data = csv_buffer.getvalue()

    filename_stem = Path(st.session_state.get('uploaded_filename', 'processed_data')).stem
    csv_filename = f"{filename_stem}_processed.csv"

    st.download_button(
        label="Download as CSV",
        data=csv_data,
        file_name=csv_filename,
        mime="text/csv",
        use_container_width=True
    )

with col2:
    st.subheader("Excel Format")
    st.caption("Professional summary with statistics")
    
    excel_data = build_excel_export(
        df_processed,
        animal_col=st.session_state.get('sex_col_used'),
        threshold=st.session_state.get('sex_threshold_used', 16),
        rebuild_sex_labels=st.session_state.get('sex_rebuild_from_threshold', True),
        include_branding=st.session_state.get('branding_enabled', False),
    )

    filename_stem = Path(st.session_state.get('uploaded_filename', 'processed_data')).stem
    excel_filename = f"{filename_stem}_processed.xlsx"

    st.download_button(
        label="Download as Excel",
        data=excel_data,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.divider()

st.subheader("Data Summary")

col1, col2, col3 = st.columns(3)
with col1:
    st.metric("Total Rows", len(df_processed))
with col2:
    st.metric("Total Columns", len(df_processed.columns))
with col3:
    st.metric("File Size (CSV)", f"{len(csv_data) / 1024:.1f} KB")

st.divider()
st.success("Your data is ready for download. Both formats contain all processed data and computed statistics.")
