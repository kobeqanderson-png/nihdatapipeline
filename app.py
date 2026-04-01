"""
Data Processing Pipeline - Interactive Web App

Run with: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
import io
import sys
import re
from scipy import stats as scipy_stats
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="NIH SABV Compliant Pipeline",
    page_icon=None,
    layout="wide"
)


def apply_custom_theme() -> None:
    """Apply a dark, monospace-first visual style across Streamlit widgets."""
    st.markdown(
        """
        <style>
        :root {
            --bg-main: #0d1117;
            --bg-panel: #161b22;
            --bg-soft: #21262d;
            --text-main: #d1d5db;
            --text-muted: #9aa4b2;
            --accent: #22c55e;
            --accent-soft: #15803d;
            --accent-deep: #166534;
            --border: #30363d;
        }

        html, body, [class*="css"] {
            font-family: "IBM Plex Mono", "JetBrains Mono", "Fira Code", monospace !important;
            background: radial-gradient(circle at top right, #1f2937 0%, var(--bg-main) 45%) !important;
            color: var(--text-main) !important;
        }

        .stApp {
            background: radial-gradient(circle at top right, #1f2937 0%, var(--bg-main) 45%) !important;
            color: var(--text-main) !important;
        }

        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #111827 0%, #0f172a 100%) !important;
            border-right: 1px solid var(--border);
        }

        .stMarkdown, .stText, .stCaption, label, p, h1, h2, h3, h4 {
            color: var(--text-main) !important;
        }

        h1, h2, h3 {
            letter-spacing: 0.02em;
            text-shadow: 0 0 0.5px rgba(34, 197, 94, 0.45);
        }

        a {
            color: var(--accent) !important;
        }

        .stButton > button,
        .stDownloadButton > button {
            background: var(--accent-soft) !important;
            color: #e6fffa !important;
            border: 1px solid var(--accent) !important;
            border-radius: 8px;
            box-shadow: 0 0 0 1px rgba(34, 197, 94, 0.25) inset;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            background: var(--accent) !important;
            color: #0d1117 !important;
            border-color: #86efac !important;
        }

        .stButton > button:focus,
        .stDownloadButton > button:focus,
        .stButton > button:focus-visible,
        .stDownloadButton > button:focus-visible {
            outline: 2px solid #86efac !important;
            outline-offset: 2px !important;
            box-shadow: none !important;
        }

        div[data-baseweb="select"] > div,
        .stTextInput > div > div,
        .stNumberInput > div > div,
        .stTextArea > div > div,
        .stDateInput > div > div {
            background-color: var(--bg-soft) !important;
            border: 1px solid var(--border) !important;
            color: var(--text-main) !important;
        }

        div[data-baseweb="select"] > div:focus-within,
        .stTextInput > div > div:focus-within,
        .stNumberInput > div > div:focus-within,
        .stTextArea > div > div:focus-within,
        .stDateInput > div > div:focus-within {
            border-color: var(--accent) !important;
            box-shadow: 0 0 0 1px rgba(34, 197, 94, 0.35) !important;
        }

        div[data-baseweb="slider"] [role="slider"] {
            background-color: var(--accent) !important;
        }

        div[data-baseweb="slider"] > div > div {
            background-color: rgba(34, 197, 94, 0.25) !important;
        }

        .stCheckbox input:checked + div,
        .stRadio input:checked + div {
            background-color: var(--accent) !important;
            border-color: var(--accent) !important;
        }

        button[role="tab"][aria-selected="true"] {
            color: #dcfce7 !important;
            border-bottom: 2px solid var(--accent) !important;
        }

        .stDataFrame, [data-testid="stTable"] {
            border: 1px solid var(--border);
            border-radius: 10px;
            overflow: hidden;
        }

        .stAlert {
            background: var(--bg-panel) !important;
            border: 1px solid var(--accent-deep) !important;
        }

        hr {
            border-color: var(--border) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


apply_custom_theme()

   
# Add project root to path for imports
project_root = Path(__file__).parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from src.data_load import read_csv, read_excel
from src.cleaning import basic_clean, save_clean
from src.features import add_log_feature
from src.visualize import boxplot_by_category, countplot

import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import train_test_split

def standardize_research_data(df, source_type):
    mappings = {
        "EthoVision XT": {"Recording time": "time", "X center": "x", "Y center": "y", "Velocity": "velocity"},
        "Any-Maze": {"Time": "time", "Centre position X": "x", "Centre position Y": "y", "Speed": "velocity"},
        "Standardized CSV": {} 
    }
    
    if source_type in mappings:
        df = df.rename(columns=mappings[source_type])
    
    if 'sex' not in df.columns:
        for col in ['Sex', 'Gender', 'SABV_Group']:
            if col in df.columns:
                df = df.rename(columns={col: 'sex'})
    
    return df

def safe_percent_diff(male_mean: float, female_mean: float) -> float:
    """Return percent difference relative to female mean; NaN if denominator is zero."""
    if pd.isna(female_mean) or female_mean == 0:
        return np.nan
    return (male_mean - female_mean) / female_mean * 100


def ttest_for_groups(df: pd.DataFrame, value_col: str, group_col: str = "Sex"):
    """Compute Welch's t-test between Male and Female groups for a variable."""
    male_data = df[df[group_col] == 'Male'][value_col].dropna()
    female_data = df[df[group_col] == 'Female'][value_col].dropna()

    if len(male_data) > 1 and len(female_data) > 1:
        t_stat, p_value = scipy_stats.ttest_ind(male_data, female_data, equal_var=False)
        return t_stat, p_value, len(male_data), len(female_data)
    return np.nan, np.nan, len(male_data), len(female_data)


def effect_size_cohens_d(group1: pd.Series, group2: pd.Series) -> float:
    """Calculate Cohen's d effect size between two groups."""
    n1, n2 = len(group1), len(group2)
    var1, var2 = group1.var(ddof=1), group2.var(ddof=1)
    
    if n1 < 2 or n2 < 2:
        return np.nan
    
    pooled_std = np.sqrt(((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2))
    if pooled_std == 0:
        return np.nan
    
    return (group1.mean() - group2.mean()) / pooled_std


def pvalue_to_label(p_value: float) -> str:
    """Convert p-value to a readable significance label."""
    if pd.isna(p_value):
        return "Insufficient data"
    return "Significant" if p_value < 0.05 else "Not Significant"


def sem(series: pd.Series) -> float:
    """Calculate SEM = std(N) / sqrt(N) using sample std (ddof=1)."""
    values = series.dropna()
    n = len(values)
    if n < 2:
        return np.nan
    return values.std(ddof=1) / np.sqrt(n)


def autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    """Autosize worksheet columns based on rendered cell text length."""
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_width), max_width)


def parse_animal_number(value) -> float:
    """Extract animal ID numbers from numeric values or labels like 'rat17'/'subject_17'."""
    if pd.isna(value):
        return np.nan

    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)

    text = str(value).strip()
    if not text:
        return np.nan

    lower_text = text.lower()

    # Explicitly support common animal ID prefixes used in lab datasets.
    prefixed_match = re.search(
        r"(?:rat|subject|animal)\s*[-_#:]?\s*(\d+(?:\.\d+)?)",
        lower_text,
    )
    if prefixed_match:
        return float(prefixed_match.group(1))

    # Also handle plain numeric strings such as "17" or "17.0".
    numeric_only_match = re.fullmatch(r"\d+(?:\.\d+)?", lower_text)
    if numeric_only_match:
        return float(numeric_only_match.group(0))

    # Backward-compatible fallback for mixed identifiers that still include a number.
    match = re.search(r"\d+(?:\.\d+)?", text)
    if match:
        return float(match.group(0))
    return np.nan


def parse_animal_number_series(series: pd.Series) -> pd.Series:
    """Vectorized animal-number parsing for mixed numeric/text identifiers."""
    return series.apply(parse_animal_number)


def parse_id_list(raw_text: str):
    """Parse comma-separated IDs/ranges (e.g., '1-16, 20, 22-24') into numeric IDs."""
    values = set()
    invalid_tokens = []

    for token in raw_text.split(','):
        token = token.strip()
        if not token:
            continue

        if '-' in token:
            parts = token.split('-', 1)
            if len(parts) != 2:
                invalid_tokens.append(token)
                continue
            try:
                start = float(parts[0].strip())
                end = float(parts[1].strip())
            except ValueError:
                invalid_tokens.append(token)
                continue

            # Expand integer-like ranges; keep decimal endpoints as explicit IDs.
            if float(start).is_integer() and float(end).is_integer():
                start_i = int(start)
                end_i = int(end)
                lo, hi = (start_i, end_i) if start_i <= end_i else (end_i, start_i)
                for v in range(lo, hi + 1):
                    values.add(float(v))
            else:
                values.add(float(start))
                values.add(float(end))
            continue

        try:
            values.add(float(token))
        except ValueError:
            invalid_tokens.append(token)

    return values, invalid_tokens


def infer_animal_column(df: pd.DataFrame):
    """Infer likely animal identifier column name from common patterns."""
    normalized = {col.lower().strip(): col for col in df.columns}
    preferred = ['animal #', 'animal', 'animal number', 'animal_number', 'animal id', 'animal_id']

    for key in preferred:
        if key in normalized:
            return normalized[key]

    for col in df.columns:
        if 'animal' in col.lower():
            return col

    return None


def build_excel_export(
    df_processed: pd.DataFrame,
    animal_col: str = None,
    threshold: float = 16,
    rebuild_sex_labels: bool = True,
) -> bytes:
    """Build one digestible Excel sheet with animal-ordered data and a summary stats table."""
    excel_buffer = io.BytesIO()

    animal_col = animal_col if animal_col in df_processed.columns else infer_animal_column(df_processed)
    export_df = df_processed.copy()

    if animal_col and rebuild_sex_labels:
        animal_num = parse_animal_number_series(export_df[animal_col])
        # Rebuild classification directly from animal number to avoid stale/mismatched sex labels.
        export_df['Sex'] = np.where(
            animal_num <= threshold,
            'Male',
            np.where(animal_num > threshold, 'Female', 'Unclassified')
        )
    elif 'Sex' not in export_df.columns:
        export_df['Sex'] = 'Unclassified'

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        sheet_name = 'Summary_By_Animal'
        header_font = Font(bold=True)
        section_font = Font(bold=True, size=12)
        highlight_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
        section_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        table_header_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        ordered_df = export_df.copy()
        low_count = 0
        has_gap = False
        if animal_col:
            ordered_df['_animal_num_sort'] = parse_animal_number_series(ordered_df[animal_col])
            low = ordered_df[ordered_df['_animal_num_sort'].notna() & (ordered_df['_animal_num_sort'] <= threshold)]
            high = ordered_df[ordered_df['_animal_num_sort'].notna() & (ordered_df['_animal_num_sort'] > threshold)]
            unknown = ordered_df[ordered_df['_animal_num_sort'].isna()]

            low_count = len(low)
            has_gap = len(low) > 0 and len(high) > 0

            pieces = [low]
            if has_gap:
                spacer = pd.DataFrame([{col: np.nan for col in ordered_df.columns}])
                pieces.append(spacer)
            pieces.extend([high, unknown])
            ordered_df = pd.concat(pieces, ignore_index=True)
            ordered_df = ordered_df.drop(columns=['_animal_num_sort'])

        ordered_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]

        ws.cell(row=1, column=1, value='Summary Ordered by Animal Number').font = Font(bold=True, size=13)
        ws.cell(row=3, column=1, value='Animal-Ordered Processed Data').font = section_font
        ws.cell(row=3, column=1).fill = section_fill

        total_cols = len(ordered_df.columns)
        header_row = 3
        data_start = 4

        for c in range(1, total_cols + 1):
            ws.cell(row=header_row, column=c).font = header_font
            ws.cell(row=header_row, column=c).fill = table_header_fill
            ws.cell(row=header_row, column=c).border = thin_border

        if animal_col and has_gap:
            spacer_row = data_start + low_count
            ws.cell(
                row=spacer_row,
                column=1,
                value=f'--- Spacer: animals > {int(threshold) if float(threshold).is_integer() else threshold} begin below ---'
            ).font = header_font
            ws.cell(row=spacer_row, column=1).fill = section_fill

        for r in range(data_start, data_start + len(ordered_df)):
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).border = thin_border

        # Summary stats table (means, SEM, and t-test)
        summary_title_row = data_start + len(ordered_df) + 2
        summary_header_row = summary_title_row + 1
        summary_data_start = summary_header_row + 1

        summary_headers = [
            'Variable', 'Male N', 'Female N', 'Male Mean', 'Female Mean', 'Male SEM', 'Female SEM',
            't-stat', 'p-value', 'p<0.005'
        ]

        ws.cell(row=summary_title_row, column=1, value='Sex Summary Table (Welch t-test)').font = Font(bold=True, size=12)
        ws.cell(row=summary_title_row + 1, column=1, value='SEM = stdev/sqrt').font = header_font

        for i, col_title in enumerate(summary_headers, start=1):
            hcell = ws.cell(row=summary_header_row, column=i, value=col_title)
            hcell.font = header_font
            hcell.fill = table_header_fill
            hcell.border = thin_border

        numeric_cols = [
            col for col in export_df.select_dtypes(include=['number']).columns
            if col != animal_col
        ]

        for i, col_name in enumerate(numeric_cols):
            row = summary_data_start + i
            male_vals = export_df[export_df['Sex'] == 'Male'][col_name].dropna()
            female_vals = export_df[export_df['Sex'] == 'Female'][col_name].dropna()
            t_stat, p_value, n_male, n_female = ttest_for_groups(export_df, col_name)
            is_highlight = not pd.isna(p_value) and p_value < 0.005

            male_mean = male_vals.mean() if len(male_vals) > 0 else np.nan
            female_mean = female_vals.mean() if len(female_vals) > 0 else np.nan
            male_sem = sem(male_vals)
            female_sem = sem(female_vals)

            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=n_male)
            ws.cell(row=row, column=3, value=n_female)
            ws.cell(row=row, column=4, value=round(float(male_mean), 6) if not pd.isna(male_mean) else np.nan)
            ws.cell(row=row, column=5, value=round(float(female_mean), 6) if not pd.isna(female_mean) else np.nan)
            ws.cell(row=row, column=6, value=round(float(male_sem), 6) if not pd.isna(male_sem) else np.nan)
            ws.cell(row=row, column=7, value=round(float(female_sem), 6) if not pd.isna(female_sem) else np.nan)
            ws.cell(row=row, column=8, value=round(float(t_stat), 6) if not pd.isna(t_stat) else np.nan)
            ws.cell(row=row, column=9, value=round(float(p_value), 6) if not pd.isna(p_value) else np.nan)
            ws.cell(row=row, column=10, value='YES' if is_highlight else 'NO')

            for col_num in range(1, len(summary_headers) + 1):
                ws.cell(row=row, column=col_num).border = thin_border

            if is_highlight:
                for col_num in range(1, len(summary_headers) + 1):
                    ws.cell(row=row, column=col_num).fill = highlight_fill

        ws.freeze_panes = 'A4'
        if len(numeric_cols) > 0:
            ws.auto_filter.ref = (
                f"A{summary_header_row}:{get_column_letter(len(summary_headers))}{summary_data_start + len(numeric_cols) - 1}"
            )
        autosize_columns(ws)

    return excel_buffer.getvalue()

# Title and description
st.title(" NIH SABV Compliant Pipeline")
st.markdown("""
Upload CSV or Excel files to analyze sex differences in your data.

**Features:**
- Load CSV and Excel files
- Automatic sex classification (Animal ≤16 = Male, >16 = Female)
- Basic data cleaning (missing values, duplicates, whitespace)
- Sex differences visualization and statistics
- Feature engineering (log transformations)
- Download processed data
""")

st.divider()

# Sidebar for settings
with st.sidebar:
    st.header("Settings")

    # Sex classification settings
    st.subheader("Sex Classification")
    st.markdown("**Animal # ≤ threshold = Male, > threshold = Female**")
    sex_threshold = st.number_input("Threshold (Animal #)", value=16, min_value=1, step=1)

    st.divider()

    # Sheet selection for Excel files
    sheet_option = st.radio(
        "Excel Sheet Selection",
        ["First sheet (index 0)", "Specify sheet name"],
        index=0
    )

    if sheet_option == "Specify sheet name":
        sheet_name = st.text_input("Sheet name", value="Sheet1")
    else:
        sheet_name = 0

    st.divider()

    # Cleaning options
    st.subheader("Cleaning Options")
    fill_missing_numeric = st.checkbox("Fill missing numeric values with median", value=True)
    remove_duplicates = st.checkbox("Remove duplicate rows", value=True)
    strip_whitespace = st.checkbox("Strip whitespace from strings", value=True)

    st.divider()

    # Feature engineering options
    st.subheader("Feature Engineering")
    add_log_features = st.checkbox("Add log-transformed features", value=True)

# File upload section
st.header("1. Upload Your Data")

uploaded_file = st.file_uploader(
    "Choose a CSV or Excel file",
    type=["csv", "xlsx", "xls"],
    help="Drag and drop or click to upload"
)

# Initialize session state
if 'df_raw' not in st.session_state:
    st.session_state.df_raw = None
if 'df_processed' not in st.session_state:
    st.session_state.df_processed = None

if uploaded_file is not None:
    # Load the file
    try:
        file_extension = Path(uploaded_file.name).suffix.lower()
        
        if file_extension == '.csv':
            # This calls your custom function in src/data_load.py
            df_raw = read_csv(uploaded_file)
        elif file_extension in ['.xlsx', '.xls']:
            xl = pd.ExcelFile(uploaded_file)
            available_sheets = xl.sheet_names
            st.info(f"Available sheets: {', '.join(available_sheets)}")
            selected_sheet = st.selectbox("Select sheet to load", available_sheets, index=0)
            df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
        
        st.session_state.df_raw = df_raw
        st.success(f"Loaded **{uploaded_file.name}** - {len(df_raw):,} rows × {len(df_raw.columns)} columns")
        
    except Exception as e:
        st.error(f"Error loading file: {e}")
        st.stop()

# Data exploration section
if st.session_state.df_raw is not None:
    df_raw = st.session_state.df_raw

    st.divider()
    st.header("2. Explore Raw Data")

    # Data preview tabs
    tab1, tab2, tab3, tab4 = st.tabs(["Preview", "Statistics", "Missing Values", "Columns"])

    with tab1:
        st.subheader("Data Preview")
        n_rows = st.slider("Rows to display", 5, 50, 10)
        st.dataframe(df_raw.head(n_rows), use_container_width=True)

    with tab2:
        st.subheader("Summary Statistics")
        numeric_cols = df_raw.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            st.dataframe(df_raw[numeric_cols].describe(), use_container_width=True)
        else:
            st.warning("No numeric columns found")

    with tab3:
        st.subheader("Missing Values")
        missing = df_raw.isnull().sum()
        missing_pct = (missing / len(df_raw) * 100).round(2)
        missing_df = pd.DataFrame({
            'Column': missing.index,
            'Missing Count': missing.values,
            'Missing %': missing_pct.values
        }).sort_values('Missing Count', ascending=False)

        missing_with_values = missing_df[missing_df['Missing Count'] > 0]
        if len(missing_with_values) > 0:
            st.dataframe(missing_with_values, use_container_width=True)
        else:
            st.success("No missing values found")

    with tab4:
        st.subheader("Column Information")
        col_info = pd.DataFrame({
            'Column': df_raw.columns,
            'Type': df_raw.dtypes.astype(str).values,
            'Non-Null Count': df_raw.notna().sum().values,
            'Unique Values': df_raw.nunique().values
        })
        st.dataframe(col_info, use_container_width=True)


   

 # --- 3. Process Data Section ---
if st.session_state.df_raw is not None:
    df_raw = st.session_state.df_raw
    st.divider()
    st.header("3. Process Data")

    # Dynamic Column Selection
    all_cols = df_raw.columns.tolist()

    col_a, col_b = st.columns(2)
    with col_a:
        sex_col = st.selectbox(
            "Select Column for Sex Classification", 
            options=all_cols,
            help="Select the ID column (e.g., 1, animal_1, rat1, subject1)."
        )

    with col_b:
        classification_mode = st.radio(
            "Classification Method",
            ["Threshold split", "Manual number lists", "Female list (others Male)"],
            horizontal=True,
        )

    threshold = 16
    male_ids_text = ""
    female_ids_text = ""

    if classification_mode == "Threshold split":
        threshold = st.number_input(
            f"Threshold for {sex_col}",
            value=16,
            step=1,
            help="Animal ID <= threshold is Male; > threshold is Female.",
        )
    elif classification_mode == "Manual number lists":
        st.caption("Enter comma-separated IDs or ranges like 1-16, 20, 22-24.")
        list_col_a, list_col_b = st.columns(2)
        with list_col_a:
            male_ids_text = st.text_input("Male IDs", value="1-16")
        with list_col_b:
            female_ids_text = st.text_input("Female IDs", value="17-32")
    else:
        st.caption("Enter female IDs; any other numeric ID will be assigned Male.")
        female_ids_text = st.text_input("Female IDs", value="17-32")

    if st.button("Run Processing Pipeline", type="primary", use_container_width=True):
        with st.spinner("Processing data..."):
            try:
                # 1. Apply basic cleaning
                df_processed = basic_clean(df_raw)
                
                # 2. Apply sex classification
                sex_numeric = parse_animal_number_series(df_processed[sex_col])
                if classification_mode == "Threshold split":
                    df_processed['Sex'] = np.where(
                        sex_numeric <= threshold,
                        'Male',
                        np.where(sex_numeric > threshold, 'Female', 'Unclassified')
                    )
                elif classification_mode == "Manual number lists":
                    male_ids, male_invalid = parse_id_list(male_ids_text)
                    female_ids, female_invalid = parse_id_list(female_ids_text)

                    if male_invalid or female_invalid:
                        bad_tokens = male_invalid + female_invalid
                        raise ValueError(
                            f"Invalid ID tokens: {', '.join(bad_tokens)}. Use numbers or ranges like 1-16, 20."
                        )

                    overlap = male_ids.intersection(female_ids)
                    if overlap:
                        st.warning(
                            f"Overlapping IDs in male/female lists: {sorted(overlap)}. Male assignment takes precedence."
                        )

                    male_mask = sex_numeric.isin(male_ids)
                    female_mask = sex_numeric.isin(female_ids)
                    df_processed['Sex'] = np.select(
                        [male_mask, female_mask],
                        ['Male', 'Female'],
                        default='Unclassified',
                    )
                else:
                    female_ids, female_invalid = parse_id_list(female_ids_text)
                    if female_invalid:
                        raise ValueError(
                            f"Invalid ID tokens: {', '.join(female_invalid)}. Use numbers or ranges like 17-32, 40."
                        )

                    is_numeric_id = sex_numeric.notna()
                    female_mask = sex_numeric.isin(female_ids)
                    df_processed['Sex'] = np.select(
                        [female_mask, is_numeric_id],
                        ['Female', 'Male'],
                        default='Unclassified',
                    )

                unclassified_count = int((df_processed['Sex'] == 'Unclassified').sum())
                if unclassified_count > 0:
                    st.warning(
                        f"{unclassified_count} row(s) could not be classified because no numeric animal ID was found "
                        f"in '{sex_col}'."
                    )
                st.session_state.sex_col_used = sex_col
                st.session_state.sex_threshold_used = float(threshold)
                st.session_state.sex_rebuild_from_threshold = classification_mode == "Threshold split"
                
                # 3. Add log features if enabled
                if add_log_features:
                    numeric_cols = df_processed.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0:
                        df_processed = add_log_feature(df_processed, col=numeric_cols[0])
                
                # 4. Save results
                st.session_state.df_processed = df_processed
                st.success("Processing complete")
                
            except Exception as e:
                st.error(f"Error during processing: {e}")
else:
    st.info("Please upload a data file in Step 1 to begin processing.")
    # Visualization section
    # --- 5. Sex Differences Analysis Section ---
if st.session_state.df_processed is not None:
    df_processed = st.session_state.df_processed
    st.divider()
    st.header("5. Sex Differences Analysis")
    
    numeric_cols = df_processed.select_dtypes(include=['number']).columns.tolist()

    # Check if Sex column exists
    if 'Sex' in df_processed.columns:
        # Sex distribution overview
        st.subheader("Sex Distribution")

        sex_counts = df_processed['Sex'].value_counts()
        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("Total Samples", len(df_processed))
        with col2:
            male_count = sex_counts.get('Male', 0)
            st.metric("Males", male_count, f"{male_count/len(df_processed)*100:.1f}%")
        with col3:
            female_count = sex_counts.get('Female', 0)
            st.metric("Females", female_count, f"{female_count/len(df_processed)*100:.1f}%")

        # Sex distribution pie chart
        fig, axes = plt.subplots(1, 2, figsize=(12, 4))

        # Pie chart
        colors = ['#3498db', '#e74c3c']  # Blue for male, red for female
        axes[0].pie(sex_counts.values, labels=sex_counts.index, autopct='%1.1f%%',
                   colors=colors, startangle=90)
        axes[0].set_title('Sex Distribution')

        # Bar chart
        sex_counts.plot(kind='bar', ax=axes[1], color=colors, edgecolor='black')
        axes[1].set_title('Sex Counts')
        axes[1].set_xlabel('Sex')
        axes[1].set_ylabel('Count')
        axes[1].tick_params(axis='x', rotation=0)

        plt.tight_layout()
        st.pyplot(fig)
        plt.close()

        st.divider()

        # Sex differences comparison
        st.subheader("Compare Variables by Sex")

        if numeric_cols:
            selected_var = st.selectbox(
                "Select variable to compare between sexes",
                numeric_cols,
                key="sex_compare"
            )

            viz_col1, viz_col2 = st.columns(2)

            with viz_col1:
                # Box plot by sex
                fig, ax = plt.subplots(figsize=(8, 5))
                df_processed.boxplot(column=selected_var, by='Sex', ax=ax)
                ax.set_title(f'{selected_var} by Sex')
                ax.set_xlabel('Sex')
                ax.set_ylabel(selected_var)
                plt.suptitle('')  # Remove automatic title
                st.pyplot(fig)
                plt.close()

            with viz_col2:
                # Violin plot by sex
                fig, ax = plt.subplots(figsize=(8, 5))
                parts = ax.violinplot(
                    [df_processed[df_processed['Sex'] == 'Male'][selected_var].dropna(),
                     df_processed[df_processed['Sex'] == 'Female'][selected_var].dropna()],
                    positions=[1, 2],
                    showmeans=True,
                    showmedians=True
                )
                ax.set_xticks([1, 2])
                ax.set_xticklabels(['Male', 'Female'])
                ax.set_title(f'{selected_var} Distribution by Sex')
                ax.set_xlabel('Sex')
                ax.set_ylabel(selected_var)
                st.pyplot(fig)
                plt.close()

            # Histogram overlay by sex
            st.subheader("Distribution Overlay")
            fig, ax = plt.subplots(figsize=(10, 5))

            male_data = df_processed[df_processed['Sex'] == 'Male'][selected_var].dropna()
            female_data = df_processed[df_processed['Sex'] == 'Female'][selected_var].dropna()

            ax.hist(male_data, bins=20, alpha=0.6, label='Male', color='#3498db', edgecolor='black')
            ax.hist(female_data, bins=20, alpha=0.6, label='Female', color='#e74c3c', edgecolor='black')
            ax.set_title(f'{selected_var} Distribution by Sex')
            ax.set_xlabel(selected_var)
            ax.set_ylabel('Frequency')
            ax.legend()
            st.pyplot(fig)
            plt.close()

            # Summary statistics by sex
            st.subheader("Summary Statistics by Sex")

            stats_by_sex = df_processed.groupby('Sex')[selected_var].agg([
                'count', 'mean', 'std', 'min', 'median', 'max'
            ]).round(4)
            stats_by_sex.columns = ['Count', 'Mean', 'Std Dev', 'Min', 'Median', 'Max']
            st.dataframe(stats_by_sex, use_container_width=True)

            # Statistical test (t-test)
            t_stat, p_value, n_male, n_female = ttest_for_groups(df_processed, selected_var)

            if not pd.isna(t_stat):

                st.subheader("Statistical Test (Independent t-test)")
                stat_col1, stat_col2, stat_col3 = st.columns(3)

                with stat_col1:
                    st.metric("t-statistic", f"{t_stat:.4f}")
                with stat_col2:
                    st.metric("p-value", f"{p_value:.4f}")
                with stat_col3:
                    significance = "Significant" if p_value < 0.05 else "Not Significant"
                    st.metric("Result (α=0.05)", significance)

                if p_value < 0.05:
                    st.success(f"Significant difference found (p={p_value:.4f} < 0.05)")
                else:
                    st.info(f"No significant difference (p={p_value:.4f} >= 0.05)")
            else:
                st.warning(
                    f"Not enough non-missing values for a t-test on {selected_var}. "
                    f"Male n={n_male}, Female n={n_female}."
                )

        st.divider()

        # Multi-variable comparison
        st.subheader("Multi-Variable Sex Comparison")

        if len(numeric_cols) > 0:
            selected_vars = st.multiselect(
                "Select variables to compare",
                numeric_cols,
                default=numeric_cols[:min(3, len(numeric_cols))]
            )

            if selected_vars:
                # Mean comparison bar chart
                fig, ax = plt.subplots(figsize=(12, 6))

                male_means = df_processed[df_processed['Sex'] == 'Male'][selected_vars].mean()
                female_means = df_processed[df_processed['Sex'] == 'Female'][selected_vars].mean()

                x = np.arange(len(selected_vars))
                width = 0.35

                bars1 = ax.bar(x - width/2, male_means, width, label='Male', color='#3498db', edgecolor='black')
                bars2 = ax.bar(x + width/2, female_means, width, label='Female', color='#e74c3c', edgecolor='black')

                ax.set_xlabel('Variables')
                ax.set_ylabel('Mean Value')
                ax.set_title('Mean Comparison by Sex')
                ax.set_xticks(x)
                ax.set_xticklabels(selected_vars, rotation=45, ha='right')
                ax.legend()

                plt.tight_layout()
                st.pyplot(fig)
                plt.close()

                # Combined bar + scatter (mean with individual points)
                st.subheader("Combined Bar + Scatter")
                combo_var = st.selectbox(
                    "Select variable for combined bar + scatter",
                    selected_vars,
                    key="combo_var"
                )
                fig, ax = plt.subplots(figsize=(10, 5))
                barplot_common = {
                    'data': df_processed,
                    'x': 'Sex',
                    'y': combo_var,
                    'order': ['Male', 'Female'],
                    'estimator': np.mean,
                    'palette': ['#3498db', '#e74c3c'],
                    'alpha': 0.75,
                    'ax': ax,
                }
                try:
                    sns.barplot(errorbar=('ci', 95), **barplot_common)
                except TypeError:
                    # Older seaborn versions use `ci` instead of `errorbar`.
                    sns.barplot(ci=95, **barplot_common)
                sns.stripplot(
                    data=df_processed,
                    x='Sex',
                    y=combo_var,
                    order=['Male', 'Female'],
                    color='black',
                    alpha=0.5,
                    jitter=0.18,
                    size=4,
                    ax=ax
                )
                ax.set_title(f'{combo_var}: Mean Bar with Sample Scatter')
                ax.set_xlabel('Sex')
                ax.set_ylabel(combo_var)
                st.pyplot(fig)
                plt.close()

                # Density distribution (KDE)
                st.subheader("Density Distribution")
                density_var = st.selectbox(
                    "Select variable for density distribution",
                    selected_vars,
                    key="density_var"
                )
                fig, ax = plt.subplots(figsize=(10, 5))
                male_density = df_processed[df_processed['Sex'] == 'Male'][density_var].dropna()
                female_density = df_processed[df_processed['Sex'] == 'Female'][density_var].dropna()

                if len(male_density) > 1:
                    sns.kdeplot(male_density, fill=True, alpha=0.4, label='Male', color='#3498db', ax=ax)
                if len(female_density) > 1:
                    sns.kdeplot(female_density, fill=True, alpha=0.4, label='Female', color='#e74c3c', ax=ax)

                ax.set_title(f'{density_var} Density by Sex')
                ax.set_xlabel(density_var)
                ax.set_ylabel('Density')
                if len(male_density) > 1 or len(female_density) > 1:
                    ax.legend()
                else:
                    st.info(
                        f"Not enough values to estimate density for {density_var}. "
                        "Need at least 2 non-missing values per group."
                    )
                st.pyplot(fig)
                plt.close()

                # Scatter plot
                st.subheader("Scatter Graph")
                scatter_cols = st.columns(2)
                with scatter_cols[0]:
                    x_var = st.selectbox("X-axis variable", selected_vars, key="scatter_x")
                with scatter_cols[1]:
                    y_candidates = [c for c in selected_vars if c != x_var] or selected_vars
                    y_var = st.selectbox("Y-axis variable", y_candidates, key="scatter_y")

                fig, ax = plt.subplots(figsize=(10, 5))
                sns.scatterplot(
                    data=df_processed,
                    x=x_var,
                    y=y_var,
                    hue='Sex',
                    palette={'Male': '#3498db', 'Female': '#e74c3c'},
                    alpha=0.75,
                    s=60,
                    ax=ax
                )
                ax.set_title(f'{y_var} vs {x_var} by Sex')
                ax.set_xlabel(x_var)
                ax.set_ylabel(y_var)
                ax.legend(title='Sex')
                st.pyplot(fig)
                plt.close()

                # Summary table
                summary_rows = []
                for var in selected_vars:
                    male_vals = df_processed[df_processed['Sex'] == 'Male'][var].dropna()
                    female_vals = df_processed[df_processed['Sex'] == 'Female'][var].dropna()
                    male_mean = male_vals.mean()
                    female_mean = female_vals.mean()
                    t_stat, p_value, n_male, n_female = ttest_for_groups(df_processed, var)

                    summary_rows.append({
                        'Variable': var,
                        'Male Mean': male_mean,
                        'Female Mean': female_mean,
                        'Difference': male_mean - female_mean,
                        'Diff %': safe_percent_diff(male_mean, female_mean),
                        'Male n': n_male,
                        'Female n': n_female,
                        't-stat': t_stat,
                        'p-value': p_value,
                        'Significance (alpha=0.05)': pvalue_to_label(p_value)
                    })

                comparison_df = pd.DataFrame({
                    'Variable': [row['Variable'] for row in summary_rows],
                    'Male Mean': [row['Male Mean'] for row in summary_rows],
                    'Female Mean': [row['Female Mean'] for row in summary_rows],
                    'Difference': [row['Difference'] for row in summary_rows],
                    'Diff %': [row['Diff %'] for row in summary_rows],
                    'Male n': [row['Male n'] for row in summary_rows],
                    'Female n': [row['Female n'] for row in summary_rows],
                    't-stat': [row['t-stat'] for row in summary_rows],
                    'p-value': [row['p-value'] for row in summary_rows],
                    'Significance (alpha=0.05)': [row['Significance (alpha=0.05)'] for row in summary_rows],
                }).round(4)
                st.dataframe(comparison_df, use_container_width=True)

        st.subheader("Effect Size Analysis")
        if selected_vars and len(selected_vars) > 0:
            effect_col1, effect_col2 = st.columns(2)
            
            with effect_col1:
                effect_var = st.selectbox("Select variable for effect size", selected_vars, key="effect_size")
            
            with effect_col2:
                st.write("")
                st.write("")
                if st.button("Calculate Cohen's d", use_container_width=True):
                    male_data = df_processed[df_processed['Sex'] == 'Male'][effect_var].dropna()
                    female_data = df_processed[df_processed['Sex'] == 'Female'][effect_var].dropna()
                    
                    if len(male_data) > 1 and len(female_data) > 1:
                        cohens_d = effect_size_cohens_d(male_data, female_data)
                        if not pd.isna(cohens_d):
                            st.metric("Cohen's d", f"{cohens_d:.4f}")
                            if abs(cohens_d) < 0.2:
                                effect_label = "Negligible"
                            elif abs(cohens_d) < 0.5:
                                effect_label = "Small"
                            elif abs(cohens_d) < 0.8:
                                effect_label = "Medium"
                            else:
                                effect_label = "Large"
                            st.info(f"Effect Size: {effect_label}")
                        else:
                            st.warning("Cannot calculate Cohen's d with this data")
                    else:
                        st.warning("Insufficient data for effect size calculation")

    else:
        st.warning("No 'Sex' column found. Please run the processing pipeline first with an Animal # column selected.")

    st.divider()
    st.header("6. General Visualizations")

    viz_col1, viz_col2 = st.columns(2)

    with viz_col1:
        st.subheader("Histogram")
        if numeric_cols:
            hist_col = st.selectbox("Select column for histogram", numeric_cols, key="hist")

            fig, ax = plt.subplots(figsize=(8, 4))
            df_processed[hist_col].dropna().hist(bins=30, ax=ax, edgecolor='black', alpha=0.7)
            ax.set_title(f'Distribution of {hist_col}')
            ax.set_xlabel(hist_col)
            ax.set_ylabel('Frequency')
            st.pyplot(fig)
            plt.close()

    with viz_col2:
        st.subheader("Box Plot")
        if numeric_cols:
            box_col = st.selectbox("Select column for box plot", numeric_cols, key="box")

            fig, ax = plt.subplots(figsize=(8, 4))
            df_processed.boxplot(column=box_col, ax=ax)
            ax.set_title(f'Box Plot of {box_col}')
            st.pyplot(fig)
            plt.close()

    # Correlation heatmap
    st.subheader("Correlation Heatmap")
    numeric_df = df_processed.select_dtypes(include=['number'])
    if len(numeric_df.columns) > 1:
        # Limit columns for readability
        max_cols = st.slider("Max columns to show", 5, 20, 10)
        cols_to_plot = numeric_df.columns[:max_cols]

        fig, ax = plt.subplots(figsize=(12, 8))
        corr_matrix = numeric_df[cols_to_plot].corr()
        sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', center=0, fmt='.2f', ax=ax)
        ax.set_title('Correlation Matrix')
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()
    else:
        st.warning("Not enough numeric columns for correlation analysis")

    st.divider()
    st.header("7. Linear Regression Modeling")

    if len(numeric_cols) < 2:
        st.warning("Linear regression requires at least two numeric columns.")
    else:
        reg_col1, reg_col2 = st.columns(2)

        with reg_col1:
            target_col = st.selectbox(
                "Target variable",
                numeric_cols,
                key="reg_target",
            )

        with reg_col2:
            feature_candidates = [c for c in numeric_cols if c != target_col]
            default_features = feature_candidates[: min(3, len(feature_candidates))]
            feature_cols = st.multiselect(
                "Feature variables",
                feature_candidates,
                default=default_features,
                key="reg_features",
            )

        model_col1, model_col2, model_col3 = st.columns(3)
        with model_col1:
            test_size = st.slider("Test split", min_value=0.1, max_value=0.4, value=0.2, step=0.05)
        with model_col2:
            random_state = st.number_input("Random state", min_value=0, value=42, step=1)
        with model_col3:
            fit_intercept = st.checkbox("Fit intercept", value=True)

        poly_col1, poly_col2 = st.columns(2)
        with poly_col1:
            use_poly = st.checkbox("Add polynomial features", value=False, key="use_poly_features")
        with poly_col2:
            poly_degree = st.slider("Polynomial degree", 2, 4, 2, disabled=not use_poly) if use_poly else 2

        if st.button("Run Linear Regression", use_container_width=True):
            if not feature_cols:
                st.warning("Select at least one feature variable.")
            else:
                model_df = df_processed[[target_col] + feature_cols].dropna()

                if len(model_df) < 10:
                    st.warning("Not enough complete rows for a stable train/test regression split. Need at least 10 rows.")
                else:
                    X = model_df[feature_cols].copy()
                    y = model_df[target_col]

                    if use_poly:
                        for col in feature_cols:
                            for d in range(2, poly_degree + 1):
                                X[f"{col}_pow{d}"] = np.power(X[col], d)

                    try:
                        X_train, X_test, y_train, y_test = train_test_split(
                            X,
                            y,
                            test_size=test_size,
                            random_state=int(random_state),
                        )

                        model = LinearRegression(fit_intercept=fit_intercept)
                        model.fit(X_train, y_train)
                        y_pred_train = model.predict(X_train)
                        y_pred_test = model.predict(X_test)

                        mae = mean_absolute_error(y_test, y_pred_test)
                        rmse = np.sqrt(mean_squared_error(y_test, y_pred_test))
                        r2 = r2_score(y_test, y_pred_test)
                        train_r2 = r2_score(y_train, y_pred_train)

                        st.subheader("Model Performance")
                        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
                        with metric_col1:
                            st.metric("R-squared (Test)", f"{r2:.4f}")
                        with metric_col2:
                            st.metric("R-squared (Train)", f"{train_r2:.4f}")
                        with metric_col3:
                            st.metric("MAE", f"{mae:.4f}")
                        with metric_col4:
                            st.metric("RMSE", f"{rmse:.4f}")

                        coef_df = pd.DataFrame({
                            "Feature": X.columns,
                            "Coefficient": model.coef_,
                        }).sort_values("Coefficient", key=np.abs, ascending=False)

                        st.subheader("Model Coefficients")
                        st.dataframe(coef_df, use_container_width=True)

                        col_coef1, col_coef2 = st.columns(2)
                        with col_coef1:
                            st.subheader("Coefficient Importance")
                            fig, ax = plt.subplots(figsize=(8, 6))
                            colors = [color for color in ["#22c55e" if x >= 0 else "#ef4444" for x in coef_df["Coefficient"]]]
                            ax.barh(coef_df["Feature"], np.abs(coef_df["Coefficient"]), color=colors, edgecolor="black")
                            ax.set_xlabel("Absolute Coefficient Value")
                            ax.set_title("Feature Importance (|Coefficient|)")
                            ax.invert_yaxis()
                            st.pyplot(fig)
                            plt.close()

                        with col_coef2:
                            st.subheader("Predicted vs Actual")
                            fig, ax = plt.subplots(figsize=(8, 6))
                            ax.scatter(y_test, y_pred_test, alpha=0.7, edgecolor="black")
                            y_min = min(float(np.min(y_test)), float(np.min(y_pred_test)))
                            y_max = max(float(np.max(y_test)), float(np.max(y_pred_test)))
                            ax.plot([y_min, y_max], [y_min, y_max], linestyle="--", color="gray")
                            ax.set_xlabel("Actual")
                            ax.set_ylabel("Predicted")
                            ax.set_title("Predicted vs Actual")
                            st.pyplot(fig)
                            plt.close()

                        st.subheader("Residual Analysis")
                        residuals = y_test - y_pred_test
                        res_col1, res_col2 = st.columns(2)
                        with res_col1:
                            fig, ax = plt.subplots(figsize=(8, 6))
                            ax.scatter(y_pred_test, residuals, alpha=0.7, edgecolor="black")
                            ax.axhline(y=0, linestyle="--", color="gray")
                            ax.set_xlabel("Predicted Values")
                            ax.set_ylabel("Residuals")
                            ax.set_title("Residual Plot")
                            st.pyplot(fig)
                            plt.close()

                        with res_col2:
                            fig, ax = plt.subplots(figsize=(8, 6))
                            ax.hist(residuals, bins=15, edgecolor="black", alpha=0.7, color="#22c55e")
                            ax.set_xlabel("Residuals")
                            ax.set_ylabel("Frequency")
                            ax.set_title("Residual Distribution")
                            st.pyplot(fig)
                            plt.close()

                    except ValueError as exc:
                        st.error(f"Linear regression failed: {exc}")

    # Download section
    st.divider()
    st.header("8. Download Processed Data")

    col1, col2 = st.columns(2)

    with col1:
        # CSV download
        csv_buffer = io.StringIO()
        df_processed.to_csv(csv_buffer, index=False)
        csv_data = csv_buffer.getvalue()

        if uploaded_file:
            csv_filename = Path(uploaded_file.name).stem + "_processed.csv"
        else:
            csv_filename = "processed_data.csv"

        st.download_button(
            label="Download as CSV",
            data=csv_data,
            file_name=csv_filename,
            mime="text/csv",
            use_container_width=True
        )

    with col2:
        # Excel download
        excel_data = build_excel_export(
            df_processed,
            animal_col=st.session_state.get('sex_col_used'),
            threshold=st.session_state.get('sex_threshold_used', 16),
            rebuild_sex_labels=st.session_state.get('sex_rebuild_from_threshold', True),
        )

        if uploaded_file:
            excel_filename = Path(uploaded_file.name).stem + "_processed.xlsx"
        else:
            excel_filename = "processed_data.xlsx"

        st.download_button(
            label="Download as Excel",
            data=excel_data,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Footer
st.divider()
st.markdown("""
---
**Data Processing Pipeline** | Built with Streamlit  
To run from command line: `streamlit run app.py`
""")








